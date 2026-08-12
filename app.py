from __future__ import annotations

from datetime import date, timedelta
from io import BytesIO
from pathlib import Path

import numpy as np
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Font, PatternFill
import pandas as pd
import plotly.express as px
import requests
import streamlit as st
import yfinance as yf


INDEX_CONFIGS = {
    "Nifty 50": {
        "url": "https://nsearchives.nseindia.com/content/indices/ind_nifty50list.csv",
        "cache": Path("data") / "latest_price_data_nifty50.csv",
        "fallback_symbols": [
            "ADANIENT",
            "ADANIPORTS",
            "APOLLOHOSP",
            "ASIANPAINT",
            "AXISBANK",
            "BAJAJ-AUTO",
            "BAJFINANCE",
            "BAJAJFINSV",
            "BEL",
            "BHARTIARTL",
            "CIPLA",
            "COALINDIA",
            "DRREDDY",
            "EICHERMOT",
            "ETERNAL",
            "GRASIM",
            "HCLTECH",
            "HDFCBANK",
            "HDFCLIFE",
            "HEROMOTOCO",
            "HINDALCO",
            "HINDUNILVR",
            "ICICIBANK",
            "INDUSINDBK",
            "INFY",
            "ITC",
            "JIOFIN",
            "JSWSTEEL",
            "KOTAKBANK",
            "LT",
            "M&M",
            "MARUTI",
            "NESTLEIND",
            "NTPC",
            "ONGC",
            "POWERGRID",
            "RELIANCE",
            "SBILIFE",
            "SBIN",
            "SHRIRAMFIN",
            "SUNPHARMA",
            "TATACONSUM",
            "TATAMOTORS",
            "TATASTEEL",
            "TCS",
            "TECHM",
            "TITAN",
            "TRENT",
            "ULTRACEMCO",
            "WIPRO",
        ],
    },
    "Nifty 500": {
        "url": "https://nsearchives.nseindia.com/content/indices/ind_nifty500list.csv",
        "cache": Path("data") / "latest_price_data_nifty500.csv",
        "fallback_symbols": [],
    },
}

FALLBACK_NIFTY50_SYMBOLS = [
    "ADANIENT",
    "ADANIPORTS",
    "APOLLOHOSP",
    "ASIANPAINT",
    "AXISBANK",
    "BAJAJ-AUTO",
    "BAJFINANCE",
    "BAJAJFINSV",
    "BEL",
    "BHARTIARTL",
    "CIPLA",
    "COALINDIA",
    "DRREDDY",
    "EICHERMOT",
    "ETERNAL",
    "GRASIM",
    "HCLTECH",
    "HDFCBANK",
    "HDFCLIFE",
    "HEROMOTOCO",
    "HINDALCO",
    "HINDUNILVR",
    "ICICIBANK",
    "INDUSINDBK",
    "INFY",
    "ITC",
    "JIOFIN",
    "JSWSTEEL",
    "KOTAKBANK",
    "LT",
    "M&M",
    "MARUTI",
    "NESTLEIND",
    "NTPC",
    "ONGC",
    "POWERGRID",
    "RELIANCE",
    "SBILIFE",
    "SBIN",
    "SHRIRAMFIN",
    "SUNPHARMA",
    "TATACONSUM",
    "TATAMOTORS",
    "TATASTEEL",
    "TCS",
    "TECHM",
    "TITAN",
    "TRENT",
    "ULTRACEMCO",
    "WIPRO",
]


REQUIRED_UPLOAD_COLUMNS = {"Date", "Stock", "Open", "High", "Low", "Close", "Volume"}
LEGACY_CACHE_PATH = Path("data") / "latest_price_data.csv"
BENCHMARK_SYMBOL = "^NSEI"
THRESHOLD_TESTS = [3.0, 4.0, 5.0, 6.0]
BACKTEST_STRATEGIES = {
    "Bullish continuation long": "Buy after bullish signals and hold for 5 trading days.",
    "Bearish bounce long": "Buy after bearish signals and hold for a 5-day bounce.",
    "Bearish continuation short": "Short after bearish signals and hold for 5 trading days.",
    "Combined long signals": "Combine bullish continuation longs and bearish bounce longs.",
}
LIMITATIONS = [
    "The study is based on historical daily data, so future market behavior can differ.",
    "Transaction costs, slippage, taxes, and liquidity constraints are not included.",
    "The analysis is exploratory and does not predict future prices.",
    "A high average return can be unreliable if it is based on very few signals.",
    "Market regime changes can affect whether momentum or reversal behavior continues.",
]


st.set_page_config(
    page_title="Nifty 50 Streak Analyzer",
    page_icon=":chart_with_upwards_trend:",
    layout="wide",
)


def add_ns_suffix(symbol: str) -> str:
    symbol = str(symbol).strip().upper()
    if not symbol:
        return symbol
    return symbol if symbol.endswith(".NS") else f"{symbol}.NS"


def get_cache_path(universe: str) -> Path:
    return INDEX_CONFIGS.get(universe, INDEX_CONFIGS["Nifty 50"])["cache"]


@st.cache_data(ttl=60 * 60 * 24)
def load_index_symbols(universe: str) -> pd.DataFrame:
    config = INDEX_CONFIGS.get(universe, INDEX_CONFIGS["Nifty 50"])
    headers = {"User-Agent": "Mozilla/5.0"}
    try:
        response = requests.get(config["url"], headers=headers, timeout=20)
        response.raise_for_status()
        frame = pd.read_csv(BytesIO(response.content))
        if "Symbol" not in frame.columns:
            raise ValueError("NSE file did not contain a Symbol column.")
        frame["Yahoo Symbol"] = frame["Symbol"].map(add_ns_suffix)
        return frame
    except Exception:
        fallback_symbols = config["fallback_symbols"]
        if not fallback_symbols:
            return pd.DataFrame(columns=["Company Name", "Industry", "Symbol", "Yahoo Symbol"])
        return pd.DataFrame(
            {
                "Company Name": fallback_symbols,
                "Industry": f"{universe} fallback list",
                "Symbol": fallback_symbols,
                "Yahoo Symbol": [add_ns_suffix(symbol) for symbol in fallback_symbols],
            }
        )


def load_nifty50_symbols() -> pd.DataFrame:
    return load_index_symbols("Nifty 50")


def normalize_yfinance_frame(raw_frame: pd.DataFrame, symbol: str) -> pd.DataFrame:
    if raw_frame.empty:
        return pd.DataFrame()

    frame = raw_frame.copy()
    if isinstance(frame.columns, pd.MultiIndex):
        first_level = frame.columns.get_level_values(0).astype(str)
        second_level = frame.columns.get_level_values(1).astype(str)
        if symbol in set(first_level):
            frame = frame[symbol].copy()
        elif symbol in set(second_level):
            frame = frame.xs(symbol, axis=1, level=1).copy()
        elif "Close" in set(first_level):
            frame.columns = first_level
        elif "Close" in set(second_level):
            frame.columns = second_level
        else:
            return pd.DataFrame()

    if isinstance(frame.columns, pd.MultiIndex):
        flattened_columns = []
        for column in frame.columns:
            parts = [str(part) for part in column if str(part) not in {"", "nan", "NaT"}]
            known_parts = [
                part
                for part in parts
                if part in {"Open", "High", "Low", "Close", "Adj Close", "Volume"}
            ]
            flattened_columns.append(known_parts[0] if known_parts else parts[0])
        frame.columns = flattened_columns

    if "Close" not in frame.columns:
        return pd.DataFrame()

    frame = frame.reset_index()
    if "Date" not in frame.columns:
        first_column = frame.columns[0]
        if first_column in {"index", "Datetime"} or pd.api.types.is_datetime64_any_dtype(
            frame[first_column]
        ):
            frame = frame.rename(columns={first_column: "Date"})

    frame["Stock"] = symbol
    keep_columns = ["Date", "Stock", "Open", "High", "Low", "Close", "Volume"]
    if not {"Date", "Stock", "Close"}.issubset(frame.columns):
        return pd.DataFrame()
    existing_columns = [column for column in keep_columns if column in frame.columns]
    return frame[existing_columns].dropna(subset=["Date", "Stock", "Close"])


@st.cache_data(show_spinner=False)
def download_yfinance_data(
    symbols: tuple[str, ...],
    start: date,
    end: date,
    batch_size: int = 80,
) -> pd.DataFrame:
    unique_symbols = tuple(symbol for symbol in dict.fromkeys(symbols) if symbol)
    all_batches = []

    for offset in range(0, len(unique_symbols), batch_size):
        batch_symbols = unique_symbols[offset : offset + batch_size]
        batch_data = download_yfinance_batch(batch_symbols, start, end)
        if not batch_data.empty:
            all_batches.append(batch_data)

    if not all_batches:
        return pd.DataFrame(columns=["Date", "Stock", "Open", "High", "Low", "Close", "Volume"])

    return (
        pd.concat(all_batches, ignore_index=True)
        .drop_duplicates(subset=["Date", "Stock"])
        .sort_values(["Stock", "Date"])
    )


def download_yfinance_batch(symbols: tuple[str, ...], start: date, end: date) -> pd.DataFrame:
    downloaded = yf.download(
        list(symbols),
        start=start,
        end=end + timedelta(days=1),
        auto_adjust=False,
        group_by="ticker",
        progress=False,
        threads=True,
    )

    rows: list[pd.DataFrame] = []
    for symbol in symbols:
        stock_frame = normalize_yfinance_frame(downloaded, symbol)
        if not stock_frame.empty:
            rows.append(stock_frame)

    keep_columns = ["Date", "Stock", "Open", "High", "Low", "Close", "Volume"]
    if rows:
        data = pd.concat(rows, ignore_index=True)
        existing_columns = [column for column in keep_columns if column in data.columns]
        data = data[existing_columns].dropna(subset=["Date", "Stock", "Close"])
    else:
        data = pd.DataFrame(columns=keep_columns)

    downloaded_symbols = set(data["Stock"].astype(str))
    missing_symbols = [symbol for symbol in symbols if symbol not in downloaded_symbols]
    retry_rows = []
    for symbol in missing_symbols:
        retry = yf.download(
            symbol,
            start=start,
            end=end + timedelta(days=1),
            auto_adjust=False,
            progress=False,
            threads=False,
        )
        retry = normalize_yfinance_frame(retry, symbol)
        if not retry.empty:
            retry_rows.append(retry)

    if retry_rows:
        retry_data = pd.concat(retry_rows, ignore_index=True)
        retry_columns = [column for column in keep_columns if column in retry_data.columns]
        if {"Date", "Stock", "Close"}.issubset(retry_columns):
            retry_data = retry_data[retry_columns].dropna(subset=["Date", "Stock", "Close"])
            data = pd.concat([data, retry_data], ignore_index=True)

    return data.drop_duplicates(subset=["Date", "Stock"]).sort_values(["Stock", "Date"])


def read_uploaded_data(uploaded_file) -> pd.DataFrame:
    if uploaded_file.name.lower().endswith((".xlsx", ".xls")):
        data = pd.read_excel(uploaded_file)
    else:
        data = pd.read_csv(uploaded_file)

    missing = REQUIRED_UPLOAD_COLUMNS.difference(data.columns)
    if missing:
        missing_text = ", ".join(sorted(missing))
        raise ValueError(f"Uploaded file is missing required columns: {missing_text}")

    data = data[list(REQUIRED_UPLOAD_COLUMNS)].copy()
    data["Date"] = pd.to_datetime(data["Date"]).dt.tz_localize(None)
    data["Stock"] = data["Stock"].astype(str).str.upper().str.strip()
    return data.dropna(subset=["Date", "Stock", "Close"])


def load_cached_price_data(cache_path: Path) -> pd.DataFrame:
    if not cache_path.exists():
        if cache_path.name == "latest_price_data_nifty50.csv" and LEGACY_CACHE_PATH.exists():
            cache_path = LEGACY_CACHE_PATH
        else:
            return pd.DataFrame()
    data = pd.read_csv(cache_path)
    data["Date"] = pd.to_datetime(data["Date"]).dt.tz_localize(None)
    return data.dropna(subset=["Date", "Stock", "Close"])


def save_cached_price_data(
    price_data: pd.DataFrame,
    benchmark_data: pd.DataFrame,
    cache_path: Path,
) -> None:
    cache_path.parent.mkdir(parents=True, exist_ok=True)
    data_to_save = pd.concat([price_data, benchmark_data], ignore_index=True)
    data_to_save.to_csv(cache_path, index=False)


def split_benchmark_data(price_data: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    data = price_data.copy()
    data["Stock"] = data["Stock"].astype(str).str.upper().str.strip()
    benchmark_mask = data["Stock"].eq(BENCHMARK_SYMBOL)
    benchmark_data = data.loc[benchmark_mask].copy()
    stock_data = data.loc[~benchmark_mask].copy()
    return stock_data, benchmark_data


def build_benchmark_forward_returns(benchmark_data: pd.DataFrame) -> pd.DataFrame:
    if benchmark_data.empty:
        return pd.DataFrame()

    benchmark = benchmark_data.sort_values("Date").copy()
    benchmark["Benchmark Next 3D Return %"] = (
        benchmark["Close"].shift(-3) / benchmark["Close"] - 1
    ) * 100
    benchmark["Benchmark Next 5D Return %"] = (
        benchmark["Close"].shift(-5) / benchmark["Close"] - 1
    ) * 100
    return benchmark[
        ["Date", "Benchmark Next 3D Return %", "Benchmark Next 5D Return %"]
    ].dropna()


def add_benchmark_columns(signals: pd.DataFrame, benchmark_data: pd.DataFrame) -> pd.DataFrame:
    if signals.empty or benchmark_data.empty:
        return signals

    benchmark_returns = build_benchmark_forward_returns(benchmark_data)
    if benchmark_returns.empty:
        return signals

    enriched = signals.merge(benchmark_returns, on="Date", how="left")
    enriched["5D Alpha vs Nifty %"] = (
        enriched["Next 5D Return %"] - enriched["Benchmark Next 5D Return %"]
    )
    return enriched


def aggregate_signal_stats(frame: pd.DataFrame, return_column: str) -> dict[str, float]:
    if frame.empty:
        return {
            "count": 0,
            "avg": np.nan,
            "median": np.nan,
            "std": np.nan,
            "best": np.nan,
            "worst": np.nan,
            "positive_rate": np.nan,
            "alpha": np.nan,
        }

    alpha = (
        frame["5D Alpha vs Nifty %"].mean()
        if "5D Alpha vs Nifty %" in frame.columns
        else np.nan
    )
    return {
        "count": len(frame),
        "avg": frame[return_column].mean(),
        "median": frame[return_column].median(),
        "std": frame[return_column].std(),
        "best": frame[return_column].max(),
        "worst": frame[return_column].min(),
        "positive_rate": frame[return_column].gt(0).mean() * 100,
        "alpha": alpha,
    }


def analyze_streaks(
    price_data: pd.DataFrame,
    threshold_pct: float,
    require_same_direction: bool,
    benchmark_data: pd.DataFrame | None = None,
) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    signals = []

    for stock, group in price_data.groupby("Stock"):
        frame = group.sort_values("Date").copy()
        frame["Daily Return %"] = frame["Close"].pct_change() * 100
        frame["3-Day Return %"] = (frame["Close"] / frame["Close"].shift(3) - 1) * 100
        frame["Next 3D Return %"] = (frame["Close"].shift(-3) / frame["Close"] - 1) * 100
        frame["Next 5D Return %"] = (frame["Close"].shift(-5) / frame["Close"] - 1) * 100
        frame["Up Day Count"] = (
            frame["Daily Return %"].gt(0).rolling(3, min_periods=3).sum()
        )
        frame["Down Day Count"] = (
            frame["Daily Return %"].lt(0).rolling(3, min_periods=3).sum()
        )

        bullish_mask = frame["3-Day Return %"].ge(threshold_pct)
        bearish_mask = frame["3-Day Return %"].le(-threshold_pct)
        if require_same_direction:
            bullish_mask &= frame["Up Day Count"].eq(3)
            bearish_mask &= frame["Down Day Count"].eq(3)

        signal_frame = frame.loc[bullish_mask | bearish_mask].copy()
        if signal_frame.empty:
            continue

        signal_frame["Signal Type"] = np.where(
            bullish_mask.loc[signal_frame.index], "Bullish Up Streak", "Bearish Down Streak"
        )
        signal_frame["Result"] = np.select(
            [
                (signal_frame["Signal Type"].eq("Bullish Up Streak"))
                & (signal_frame["Next 5D Return %"].gt(0)),
                (signal_frame["Signal Type"].eq("Bullish Up Streak"))
                & (signal_frame["Next 5D Return %"].lt(0)),
                (signal_frame["Signal Type"].eq("Bearish Down Streak"))
                & (signal_frame["Next 5D Return %"].gt(0)),
                (signal_frame["Signal Type"].eq("Bearish Down Streak"))
                & (signal_frame["Next 5D Return %"].lt(0)),
            ],
            [
                "Continued Up",
                "Reversed Down",
                "Reversed Up / Bounce",
                "Continued Down",
            ],
            default="Flat / Not Available",
        )
        signals.append(signal_frame)

    if not signals:
        empty_columns = [
            "Date",
            "Stock",
            "Signal Type",
            "Close",
            "Daily Return %",
            "3-Day Return %",
            "Next 3D Return %",
            "Next 5D Return %",
            "Result",
        ]
        return pd.DataFrame(columns=empty_columns), pd.DataFrame(), pd.DataFrame()

    all_signals = pd.concat(signals, ignore_index=True)
    all_signals = all_signals.dropna(subset=["Next 3D Return %", "Next 5D Return %"])
    output_columns = [
        "Date",
        "Stock",
        "Signal Type",
        "Close",
        "Daily Return %",
        "3-Day Return %",
        "Next 3D Return %",
        "Next 5D Return %",
        "Result",
    ]
    all_signals = all_signals[output_columns].sort_values(["Date", "Stock"])
    bullish = all_signals[all_signals["Signal Type"].eq("Bullish Up Streak")].copy()
    bearish = all_signals[all_signals["Signal Type"].eq("Bearish Down Streak")].copy()
    if benchmark_data is not None and not benchmark_data.empty:
        bullish = add_benchmark_columns(bullish, benchmark_data)
        bearish = add_benchmark_columns(bearish, benchmark_data)
    summary = build_stock_summary(bullish, bearish)
    return bullish, bearish, summary


def build_stock_summary(bullish: pd.DataFrame, bearish: pd.DataFrame) -> pd.DataFrame:
    all_stocks = sorted(set(bullish["Stock"]).union(set(bearish["Stock"])))
    rows = []

    for stock in all_stocks:
        bull = bullish[bullish["Stock"].eq(stock)]
        bear = bearish[bearish["Stock"].eq(stock)]

        bull_stats = aggregate_signal_stats(bull, "Next 5D Return %")
        bear_stats = aggregate_signal_stats(bear, "Next 5D Return %")
        bull_count = bull_stats["count"]
        bear_count = bear_stats["count"]
        bull_avg_5d = bull_stats["avg"]
        bear_avg_5d = bear_stats["avg"]
        bull_win_rate = bull_stats["positive_rate"]
        bear_bounce_rate = bear_stats["positive_rate"]

        interpretation_parts = []
        if bull_count:
            if bull_avg_5d > 0 and bull_win_rate >= 50:
                interpretation_parts.append("continues after up streaks")
            elif bull_avg_5d < 0:
                interpretation_parts.append("reverses after up streaks")
            else:
                interpretation_parts.append("mixed after up streaks")
        if bear_count:
            if bear_avg_5d > 0 and bear_bounce_rate >= 50:
                interpretation_parts.append("bounces after down streaks")
            elif bear_avg_5d < 0:
                interpretation_parts.append("continues down after down streaks")
            else:
                interpretation_parts.append("mixed after down streaks")

        rows.append(
            {
                "Stock": stock,
                "Bullish Signal Count": bull_count,
                "Avg Next 3D After Bullish %": bull["Next 3D Return %"].mean()
                if bull_count
                else np.nan,
                "Avg Next 5D After Bullish %": bull_avg_5d,
                "Median Next 5D After Bullish %": bull_stats["median"],
                "Std Dev Next 5D After Bullish %": bull_stats["std"],
                "Best Next 5D After Bullish %": bull_stats["best"],
                "Worst Next 5D After Bullish %": bull_stats["worst"],
                "Bullish Continuation Rate %": bull_win_rate,
                "Avg 5D Alpha After Bullish %": bull_stats["alpha"],
                "Bearish Signal Count": bear_count,
                "Avg Next 3D After Bearish %": bear["Next 3D Return %"].mean()
                if bear_count
                else np.nan,
                "Avg Next 5D After Bearish %": bear_avg_5d,
                "Median Next 5D After Bearish %": bear_stats["median"],
                "Std Dev Next 5D After Bearish %": bear_stats["std"],
                "Best Next 5D After Bearish %": bear_stats["best"],
                "Worst Next 5D After Bearish %": bear_stats["worst"],
                "Bearish Bounce Rate %": bear_bounce_rate,
                "Avg 5D Alpha After Bearish %": bear_stats["alpha"],
                "Final Interpretation": "; ".join(interpretation_parts) or "no signals",
            }
        )

    summary = pd.DataFrame(rows)
    if summary.empty:
        return summary

    summary["Total Signals"] = (
        summary["Bullish Signal Count"] + summary["Bearish Signal Count"]
    )
    return summary.sort_values(
        ["Total Signals", "Avg Next 5D After Bullish %", "Avg Next 5D After Bearish %"],
        ascending=[False, False, False],
    ).drop(columns=["Total Signals"])


def percentile_score(series: pd.Series, higher_is_better: bool = True) -> pd.Series:
    numeric = pd.to_numeric(series, errors="coerce")
    if numeric.notna().sum() <= 1:
        return pd.Series(50.0, index=series.index)
    ranks = numeric.rank(pct=True) * 100
    if not higher_is_better:
        ranks = 100 - ranks
    return ranks.fillna(50)


def compute_rsi(close: pd.Series, period: int = 14) -> float:
    delta = close.diff()
    gains = delta.clip(lower=0)
    losses = -delta.clip(upper=0)
    avg_gain = gains.rolling(period, min_periods=period).mean()
    avg_loss = losses.rolling(period, min_periods=period).mean()
    rs = avg_gain / avg_loss.replace(0, np.nan)
    rsi = 100 - (100 / (1 + rs))
    latest = rsi.iloc[-1] if not rsi.empty else np.nan
    if pd.isna(latest) and avg_loss.iloc[-1] == 0 and avg_gain.iloc[-1] > 0:
        return 100.0
    return float(latest) if pd.notna(latest) else np.nan


def build_benchmark_lookback_returns(
    benchmark_data: pd.DataFrame,
    lookback_days: int,
) -> pd.DataFrame:
    if benchmark_data.empty:
        return pd.DataFrame()
    benchmark = benchmark_data.sort_values("Date").copy()
    benchmark["Benchmark Lookback Return %"] = (
        benchmark["Close"] / benchmark["Close"].shift(lookback_days) - 1
    ) * 100
    return benchmark[["Date", "Benchmark Lookback Return %"]].dropna()


def build_trend_analysis(
    price_data: pd.DataFrame,
    benchmark_data: pd.DataFrame,
    stock_summary: pd.DataFrame,
    lookback_days: int,
) -> pd.DataFrame:
    rows = []
    benchmark_returns = build_benchmark_lookback_returns(benchmark_data, lookback_days)

    for stock, group in price_data.groupby("Stock"):
        frame = group.sort_values("Date").copy()
        if len(frame) <= lookback_days:
            continue

        frame["Daily Return %"] = frame["Close"].pct_change() * 100
        latest = frame.iloc[-1]
        window = frame.tail(lookback_days + 1).copy()
        lookback_return = (window["Close"].iloc[-1] / window["Close"].iloc[0] - 1) * 100
        lookback_daily_returns = window["Daily Return %"].dropna()
        avg_daily_return = lookback_daily_returns.mean()
        volatility = lookback_daily_returns.std()
        up_days = int(lookback_daily_returns.gt(0).sum())
        down_days = int(lookback_daily_returns.lt(0).sum())
        flat_days = int(lookback_days - up_days - down_days)
        start_volume = window["Volume"].iloc[0] if "Volume" in window.columns else np.nan
        end_volume = window["Volume"].iloc[-1] if "Volume" in window.columns else np.nan
        avg_volume = window["Volume"].tail(lookback_days).mean() if "Volume" in window.columns else np.nan
        volume_growth = (
            ((end_volume / start_volume) - 1) * 100
            if pd.notna(start_volume) and start_volume != 0 and pd.notna(end_volume)
            else np.nan
        )
        moving_average = window["Close"].tail(lookback_days).mean()
        price_vs_ma = ((latest["Close"] / moving_average) - 1) * 100 if moving_average else np.nan
        rsi = compute_rsi(frame["Close"], period=min(14, lookback_days))

        benchmark_return = np.nan
        if not benchmark_returns.empty:
            matched = benchmark_returns[benchmark_returns["Date"].eq(latest["Date"])]
            if not matched.empty:
                benchmark_return = matched["Benchmark Lookback Return %"].iloc[0]

        summary_row = stock_summary[stock_summary["Stock"].eq(stock)]
        continuation_rate = (
            summary_row["Bullish Continuation Rate %"].iloc[0]
            if not summary_row.empty and "Bullish Continuation Rate %" in summary_row
            else np.nan
        )
        avg_alpha = (
            summary_row["Avg 5D Alpha After Bullish %"].iloc[0]
            if not summary_row.empty and "Avg 5D Alpha After Bullish %" in summary_row
            else np.nan
        )

        rows.append(
            {
                "Stock": stock,
                "Latest Date": latest["Date"],
                "Latest Close": latest["Close"],
                f"{lookback_days}D Return %": lookback_return,
                f"{lookback_days}D Avg Daily Return %": avg_daily_return,
                f"{lookback_days}D Volatility %": volatility,
                f"{lookback_days}D Up Days": up_days,
                f"{lookback_days}D Down Days": down_days,
                f"{lookback_days}D Flat Days": flat_days,
                f"{lookback_days}D Volume Growth %": volume_growth,
                f"{lookback_days}D Avg Volume": avg_volume,
                f"{lookback_days}D Price vs MA %": price_vs_ma,
                "RSI": rsi,
                f"Benchmark {lookback_days}D Return %": benchmark_return,
                f"Relative Strength vs Nifty %": lookback_return - benchmark_return
                if pd.notna(benchmark_return)
                else np.nan,
                "Bullish Continuation Rate %": continuation_rate,
                "Avg 5D Alpha After Bullish %": avg_alpha,
            }
        )

    trend = pd.DataFrame(rows)
    if trend.empty:
        return trend

    return_col = f"{lookback_days}D Return %"
    vol_col = f"{lookback_days}D Volatility %"
    volume_col = f"{lookback_days}D Volume Growth %"
    relative_col = f"Relative Strength vs Nifty %"

    trend["Return Component"] = percentile_score(trend[return_col])
    trend["Continuation Component"] = percentile_score(trend["Bullish Continuation Rate %"])
    trend["Alpha Component"] = percentile_score(trend["Avg 5D Alpha After Bullish %"])
    trend["Low Volatility Component"] = percentile_score(trend[vol_col], higher_is_better=False)
    trend["Volume Component"] = percentile_score(trend[volume_col])
    trend["Relative Strength Component"] = percentile_score(trend[relative_col])
    trend["Trend Score"] = (
        trend["Return Component"] * 0.30
        + trend["Continuation Component"] * 0.20
        + trend["Alpha Component"] * 0.20
        + trend["Low Volatility Component"] * 0.10
        + trend["Volume Component"] * 0.10
        + trend["Relative Strength Component"] * 0.10
    )
    trend["Trend Label"] = pd.cut(
        trend["Trend Score"],
        bins=[-np.inf, 35, 55, 75, np.inf],
        labels=["Weak", "Neutral", "Strong", "Very Strong"],
    ).astype(str)

    return trend.sort_values("Trend Score", ascending=False)


def clamp_score(value: float) -> float:
    if pd.isna(value):
        return 50.0
    return float(np.clip(value, 0, 100))


def score_from_signed_percent(value: float, scale: float = 10.0) -> float:
    if pd.isna(value):
        return 50.0
    return clamp_score(50 + (value / scale) * 50)


def classify_price_volume_day(row: pd.Series) -> str:
    price_up = row["Daily Return %"] > 0
    volume_up = row["Volume Change %"] > 0
    if price_up and volume_up:
        return "Accumulation"
    if price_up and not volume_up:
        return "Weak Rally"
    if not price_up and volume_up:
        return "Distribution"
    return "Weak Selling"


def volume_spike_label(spike_ratio: float) -> str:
    if pd.isna(spike_ratio):
        return "Not Available"
    if spike_ratio >= 3:
        return "Exceptional"
    if spike_ratio >= 2:
        return "Very High"
    if spike_ratio >= 1.5:
        return "High"
    return "Normal"


def pressure_label(buying_pressure: float, selling_pressure: float) -> str:
    if buying_pressure >= 65 and buying_pressure - selling_pressure >= 10:
        return "Buying Pressure Dominant"
    if selling_pressure >= 65 and selling_pressure - buying_pressure >= 10:
        return "Selling Pressure Dominant"
    if buying_pressure >= 55 and selling_pressure >= 55:
        return "High Two-Way Participation"
    return "Balanced / Neutral"


def liquidity_rating(avg_volume: float, turnover: float, stability_score: float) -> str:
    if pd.isna(avg_volume) or pd.isna(turnover):
        return "Not Available"
    if avg_volume >= 1_000_000 and turnover >= 1_000_000_000 and stability_score >= 45:
        return "Excellent"
    if avg_volume >= 300_000 and turnover >= 200_000_000:
        return "Good"
    if avg_volume >= 100_000:
        return "Moderate"
    return "Low"


def build_volume_intelligence(
    price_data: pd.DataFrame,
    trend_analysis: pd.DataFrame,
    lookback_days: int,
) -> pd.DataFrame:
    rows = []
    trend_lookup = (
        trend_analysis.set_index("Stock")
        if not trend_analysis.empty and "Stock" in trend_analysis.columns
        else pd.DataFrame()
    )

    for stock, group in price_data.groupby("Stock"):
        frame = group.sort_values("Date").copy()
        if len(frame) <= lookback_days:
            continue

        frame["Daily Return %"] = frame["Close"].pct_change() * 100
        frame["Volume Change %"] = frame["Volume"].pct_change() * 100
        frame["Volume MA"] = frame["Volume"].rolling(lookback_days, min_periods=lookback_days).mean()

        window = frame.tail(lookback_days).copy()
        latest = frame.iloc[-1]
        prior_high = frame["High"].iloc[-lookback_days - 1 : -1].max()
        current_volume = latest["Volume"]
        avg_volume = window["Volume"].mean()
        max_volume = window["Volume"].max()
        min_volume = window["Volume"].min()
        total_volume = window["Volume"].sum()
        volume_growth = (
            ((window["Volume"].iloc[-1] / window["Volume"].iloc[0]) - 1) * 100
            if window["Volume"].iloc[0] != 0
            else np.nan
        )
        spike_ratio = current_volume / avg_volume if avg_volume else np.nan
        turnover = avg_volume * latest["Close"] if pd.notna(avg_volume) else np.nan
        volume_cv = window["Volume"].std() / avg_volume if avg_volume else np.nan
        stability_score = clamp_score(100 - volume_cv * 100) if pd.notna(volume_cv) else np.nan

        clean_window = window.dropna(subset=["Daily Return %", "Volume Change %"]).copy()
        clean_window["PV Classification"] = clean_window.apply(classify_price_volume_day, axis=1)
        counts = clean_window["PV Classification"].value_counts()
        accumulation_days = int(counts.get("Accumulation", 0))
        distribution_days = int(counts.get("Distribution", 0))
        weak_rally_days = int(counts.get("Weak Rally", 0))
        weak_selling_days = int(counts.get("Weak Selling", 0))

        high_low_range = latest["High"] - latest["Low"]
        close_location = (
            ((latest["Close"] - latest["Low"]) / high_low_range) * 100
            if high_low_range
            else 50
        )
        lookback_return = (window["Close"].iloc[-1] / window["Close"].iloc[0] - 1) * 100
        up_day_ratio = clean_window["Daily Return %"].gt(0).mean() * 100 if not clean_window.empty else 50
        down_day_ratio = clean_window["Daily Return %"].lt(0).mean() * 100 if not clean_window.empty else 50
        consecutive_green = 0
        consecutive_red = 0
        for value in reversed(clean_window["Daily Return %"].tolist()):
            if value > 0 and consecutive_red == 0:
                consecutive_green += 1
            elif value < 0 and consecutive_green == 0:
                consecutive_red += 1
            else:
                break

        buying_pressure = clamp_score(
            score_from_signed_percent(lookback_return, 12) * 0.35
            + score_from_signed_percent(volume_growth, 80) * 0.25
            + close_location * 0.25
            + min(consecutive_green * 12, 100) * 0.15
        )
        selling_pressure = clamp_score(
            score_from_signed_percent(-lookback_return, 12) * 0.35
            + score_from_signed_percent(volume_growth, 80) * 0.25
            + (100 - close_location) * 0.25
            + min(consecutive_red * 12, 100) * 0.15
        )

        mfm_denominator = (window["High"] - window["Low"]).replace(0, np.nan)
        money_flow_multiplier = ((window["Close"] - window["Low"]) - (window["High"] - window["Close"])) / mfm_denominator
        money_flow_volume = (money_flow_multiplier.fillna(0) * window["Volume"]).sum()
        adl_normalized = (money_flow_volume / total_volume) * 100 if total_volume else np.nan
        accumulation_score = clamp_score((adl_normalized + 100) / 2) if pd.notna(adl_normalized) else np.nan

        price_breakout = bool(pd.notna(prior_high) and latest["Close"] > prior_high)
        if price_breakout and spike_ratio >= 1.5:
            breakout_status = "Confirmed Breakout"
        elif price_breakout:
            breakout_status = "Weak Breakout"
        else:
            breakout_status = "No Breakout"

        consecutive_volume_increase = 0
        for value in reversed(clean_window["Volume Change %"].tolist()):
            if value > 0:
                consecutive_volume_increase += 1
            else:
                break

        market_interest = clamp_score(
            score_from_signed_percent(volume_growth, 80) * 0.40
            + score_from_signed_percent(lookback_return, 12) * 0.35
            + min(consecutive_volume_increase * 20, 100) * 0.25
        )

        trend_score = np.nan
        relative_strength = np.nan
        rsi = np.nan
        continuation = np.nan
        alpha = np.nan
        if not trend_lookup.empty and stock in trend_lookup.index:
            trend_row = trend_lookup.loc[stock]
            trend_score = trend_row.get("Trend Score", np.nan)
            relative_strength = trend_row.get("Relative Strength vs Nifty %", np.nan)
            rsi = trend_row.get("RSI", np.nan)
            continuation = trend_row.get("Bullish Continuation Rate %", np.nan)
            alpha = trend_row.get("Avg 5D Alpha After Bullish %", np.nan)

        volume_intelligence_score = clamp_score(
            buying_pressure * 0.25
            + accumulation_score * 0.25
            + market_interest * 0.20
            + score_from_signed_percent(volume_growth, 80) * 0.15
            + stability_score * 0.15
        )
        conviction_score = clamp_score(
            (trend_score if pd.notna(trend_score) else 50) * 0.35
            + volume_intelligence_score * 0.30
            + score_from_signed_percent(relative_strength, 10) * 0.15
            + (continuation if pd.notna(continuation) else 50) * 0.10
            + score_from_signed_percent(alpha, 5) * 0.10
        )
        institutional_activity = (
            "High"
            if spike_ratio >= 1.5 and abs(lookback_return) >= 5 and max(buying_pressure, selling_pressure) >= 65
            else "Medium"
            if spike_ratio >= 1.2 or max(buying_pressure, selling_pressure) >= 60
            else "Low"
        )

        rows.append(
            {
                "Stock": stock,
                "Latest Date": latest["Date"],
                "Latest Close": latest["Close"],
                "Current Volume": current_volume,
                f"Average {lookback_days}D Volume": avg_volume,
                f"Maximum {lookback_days}D Volume": max_volume,
                f"Minimum {lookback_days}D Volume": min_volume,
                f"Total {lookback_days}D Volume": total_volume,
                f"{lookback_days}D Volume Growth %": volume_growth,
                "Volume Spike Ratio": spike_ratio,
                "Volume Spike Category": volume_spike_label(spike_ratio),
                "Accumulation Days": accumulation_days,
                "Distribution Days": distribution_days,
                "Weak Rally Days": weak_rally_days,
                "Weak Selling Days": weak_selling_days,
                "Price-Volume Interpretation": pressure_label(buying_pressure, selling_pressure),
                "Buying Pressure Index": buying_pressure,
                "Selling Pressure Index": selling_pressure,
                "Accumulation Score": accumulation_score,
                "Market Interest Index": market_interest,
                "Breakout Status": breakout_status,
                "Institutional Activity Proxy": institutional_activity,
                "Liquidity Rating": liquidity_rating(avg_volume, turnover, stability_score),
                "Average Turnover": turnover,
                "Volume Stability Score": stability_score,
                "Trend Score": trend_score,
                "Volume Intelligence Score": volume_intelligence_score,
                "Conviction Score": conviction_score,
                "AI Interpretation": (
                    f"{stock} shows {pressure_label(buying_pressure, selling_pressure).lower()} over the last "
                    f"{lookback_days} sessions. Volume is {volume_spike_label(spike_ratio).lower()} versus its "
                    f"lookback average, with {accumulation_days} accumulation days and {distribution_days} "
                    f"distribution days. Breakout reading: {breakout_status.lower()}. Overall conviction score is "
                    f"{conviction_score:.1f}/100."
                ),
            }
        )

    volume = pd.DataFrame(rows)
    if volume.empty:
        return volume
    return volume.sort_values("Conviction Score", ascending=False)


def build_volume_spike_reactions(
    price_data: pd.DataFrame,
    lookback_days: int,
    spike_multiplier: float,
) -> pd.DataFrame:
    rows = []
    for stock, group in price_data.groupby("Stock"):
        frame = group.sort_values("Date").copy()
        if len(frame) <= lookback_days + 5:
            continue
        frame["Daily Return %"] = frame["Close"].pct_change() * 100
        frame["Rolling Avg Volume"] = frame["Volume"].rolling(lookback_days, min_periods=lookback_days).mean().shift(1)
        frame["Volume Spike Ratio"] = frame["Volume"] / frame["Rolling Avg Volume"]
        frame["Next 1D Return %"] = (frame["Close"].shift(-1) / frame["Close"] - 1) * 100
        frame["Next 3D Return %"] = (frame["Close"].shift(-3) / frame["Close"] - 1) * 100
        frame["Next 5D Return %"] = (frame["Close"].shift(-5) / frame["Close"] - 1) * 100
        spike_rows = frame[
            frame["Volume Spike Ratio"].ge(spike_multiplier)
            & frame["Next 5D Return %"].notna()
        ].copy()
        if spike_rows.empty:
            continue
        spike_rows["Stock"] = stock
        spike_rows["Spike Category"] = spike_rows["Volume Spike Ratio"].map(volume_spike_label)
        rows.append(
            spike_rows[
                [
                    "Date",
                    "Stock",
                    "Close",
                    "Volume",
                    "Rolling Avg Volume",
                    "Volume Spike Ratio",
                    "Spike Category",
                    "Daily Return %",
                    "Next 1D Return %",
                    "Next 3D Return %",
                    "Next 5D Return %",
                ]
            ]
        )

    if not rows:
        return pd.DataFrame()
    return pd.concat(rows, ignore_index=True).sort_values(["Date", "Stock"])


def classify_volume_regime(volume_ratio: float) -> str:
    if pd.isna(volume_ratio):
        return "Not Available"
    if volume_ratio < 0.5:
        return "Extremely Low"
    if volume_ratio < 0.8:
        return "Low"
    if volume_ratio < 1.2:
        return "Average"
    if volume_ratio < 1.5:
        return "High"
    if volume_ratio < 2:
        return "Very High"
    return "Exceptional"


def classify_trend_regime(return_20d: float) -> str:
    if pd.isna(return_20d):
        return "Not Available"
    if return_20d >= 3:
        return "Bullish"
    if return_20d <= -3:
        return "Bearish"
    return "Neutral"


def classify_outlook(probability_positive: float, expected_return: float) -> str:
    if pd.isna(probability_positive) or pd.isna(expected_return):
        return "Insufficient Evidence"
    if probability_positive >= 65 and expected_return > 0:
        return "Bullish"
    if probability_positive <= 35 and expected_return < 0:
        return "Bearish"
    if expected_return > 0:
        return "Moderately Bullish"
    if expected_return < 0:
        return "Moderately Bearish"
    return "Neutral"


def classify_confidence(sample_size: int, probability_positive: float, std_dev: float) -> tuple[float, str]:
    sample_score = min(sample_size / 100 * 55, 55)
    consistency_score = abs(probability_positive - 50) * 0.5 if pd.notna(probability_positive) else 0
    stability_score = clamp_score(100 - std_dev * 8) * 0.20 if pd.notna(std_dev) else 10
    confidence = clamp_score(sample_score + consistency_score + stability_score)
    if confidence >= 75:
        return confidence, "High"
    if confidence >= 50:
        return confidence, "Medium"
    return confidence, "Low"


def classify_risk(worst_loss: float, var_95: float, std_dev: float) -> str:
    risk_value = np.nanmean([abs(worst_loss), abs(var_95), std_dev])
    if pd.isna(risk_value):
        return "Not Available"
    if risk_value >= 8:
        return "High"
    if risk_value >= 4:
        return "Moderate"
    return "Low"


def future_window_extreme(series: pd.Series, horizon: int, mode: str) -> pd.Series:
    shifted = series.shift(-1)
    reversed_window = shifted.iloc[::-1].rolling(horizon, min_periods=horizon)
    result = reversed_window.max() if mode == "max" else reversed_window.min()
    return result.iloc[::-1]


def build_long_term_volume_profile(price_data: pd.DataFrame) -> pd.DataFrame:
    rows = []
    for stock, group in price_data.groupby("Stock"):
        frame = group.sort_values("Date").copy()
        if frame.empty:
            continue
        volume = frame["Volume"].dropna()
        if volume.empty:
            continue
        latest = frame.iloc[-1]
        latest_volume = latest["Volume"]
        mean_volume = volume.mean()
        std_volume = volume.std()
        latest_ratio = latest_volume / mean_volume if mean_volume else np.nan
        volume_z_score = (latest_volume - mean_volume) / std_volume if std_volume else np.nan
        rows.append(
            {
                "Stock": stock,
                "Latest Date": latest["Date"],
                "Latest Volume": latest_volume,
                "5Y Mean Volume": mean_volume,
                "5Y Median Volume": volume.median(),
                "5Y Volume Std Dev": std_volume,
                "25th Percentile Volume": volume.quantile(0.25),
                "50th Percentile Volume": volume.quantile(0.50),
                "75th Percentile Volume": volume.quantile(0.75),
                "90th Percentile Volume": volume.quantile(0.90),
                "95th Percentile Volume": volume.quantile(0.95),
                "Latest Volume Ratio vs 5Y Mean": latest_ratio,
                "Latest Volume Z Score": volume_z_score,
                "Latest Volume Regime": classify_volume_regime(latest_ratio),
            }
        )
    profile = pd.DataFrame(rows)
    if profile.empty:
        return profile
    return profile.sort_values("Latest Volume Z Score", ascending=False)


def build_historical_behavior_database(
    price_data: pd.DataFrame,
    benchmark_data: pd.DataFrame,
) -> pd.DataFrame:
    rows = []
    benchmark = pd.DataFrame()
    if not benchmark_data.empty:
        benchmark = benchmark_data.sort_values("Date").copy()
        benchmark["Benchmark 20D Return %"] = (benchmark["Close"] / benchmark["Close"].shift(20) - 1) * 100
        benchmark = benchmark[["Date", "Benchmark 20D Return %"]]

    for stock, group in price_data.groupby("Stock"):
        frame = group.sort_values("Date").copy()
        if len(frame) < 260:
            continue
        frame["Daily Return %"] = frame["Close"].pct_change() * 100
        frame["Volume Change %"] = frame["Volume"].pct_change() * 100
        frame["20D Return %"] = (frame["Close"] / frame["Close"].shift(20) - 1) * 100
        frame["50D Volume Average"] = frame["Volume"].rolling(50, min_periods=20).mean()
        frame["100D Volume Average"] = frame["Volume"].rolling(100, min_periods=50).mean()
        frame["200D Volume Average"] = frame["Volume"].rolling(200, min_periods=100).mean()
        frame["Volume Ratio"] = frame["Volume"] / frame["50D Volume Average"]
        frame["Volume Regime"] = frame["Volume Ratio"].map(classify_volume_regime)
        frame["Trend Regime"] = frame["20D Return %"].map(classify_trend_regime)
        frame["RSI"] = frame["Close"].rolling(15, min_periods=15).apply(
            lambda values: compute_rsi(pd.Series(values), period=14),
            raw=False,
        )
        frame["Price-Volume Behaviour"] = frame.apply(classify_price_volume_day, axis=1)
        if not benchmark.empty:
            frame = frame.merge(benchmark, on="Date", how="left")
            frame["Relative Strength vs Nifty %"] = frame["20D Return %"] - frame["Benchmark 20D Return %"]
        else:
            frame["Relative Strength vs Nifty %"] = np.nan

        for horizon in [5, 10, 15]:
            frame[f"Future {horizon}D Return %"] = (frame["Close"].shift(-horizon) / frame["Close"] - 1) * 100
            future_high = future_window_extreme(frame["High"], horizon, "max")
            future_low = future_window_extreme(frame["Low"], horizon, "min")
            frame[f"Future {horizon}D Max Upside %"] = (future_high / frame["Close"] - 1) * 100
            frame[f"Future {horizon}D Max Drawdown %"] = (future_low / frame["Close"] - 1) * 100
            frame[f"Future {horizon}D Direction"] = np.where(
                frame[f"Future {horizon}D Return %"].gt(0),
                "Positive",
                "Negative",
            )

        rows.append(
            frame[
                [
                    "Date",
                    "Stock",
                    "Open",
                    "High",
                    "Low",
                    "Close",
                    "Volume",
                    "Daily Return %",
                    "20D Return %",
                    "50D Volume Average",
                    "100D Volume Average",
                    "200D Volume Average",
                    "Volume Ratio",
                    "Volume Regime",
                    "Trend Regime",
                    "RSI",
                    "Relative Strength vs Nifty %",
                    "Price-Volume Behaviour",
                    "Future 5D Return %",
                    "Future 5D Max Upside %",
                    "Future 5D Max Drawdown %",
                    "Future 5D Direction",
                    "Future 10D Return %",
                    "Future 10D Max Upside %",
                    "Future 10D Max Drawdown %",
                    "Future 10D Direction",
                    "Future 15D Return %",
                    "Future 15D Max Upside %",
                    "Future 15D Max Drawdown %",
                    "Future 15D Direction",
                ]
            ]
        )

    if not rows:
        return pd.DataFrame()
    observations = pd.concat(rows, ignore_index=True)
    return observations.dropna(subset=["Volume Ratio", "RSI", "Future 15D Return %"])


def summarize_matches(matches: pd.DataFrame, horizon: int) -> dict[str, float | str]:
    return_col = f"Future {horizon}D Return %"
    upside_col = f"Future {horizon}D Max Upside %"
    drawdown_col = f"Future {horizon}D Max Drawdown %"
    returns = matches[return_col].dropna()
    sample_size = len(returns)
    if sample_size == 0:
        return {
            "Sample Size": 0,
            "Probability Positive %": np.nan,
            "Average Return %": np.nan,
            "Median Return %": np.nan,
            "Best Historical Gain %": np.nan,
            "Worst Historical Loss %": np.nan,
            "Average Max Upside %": np.nan,
            "Average Max Drawdown %": np.nan,
            "VaR 95% %": np.nan,
            "Return Std Dev %": np.nan,
            "Confidence Score": np.nan,
            "Confidence Label": "Low",
            "Risk Label": "Not Available",
            "Outlook": "Insufficient Evidence",
        }

    probability_positive = returns.gt(0).mean() * 100
    avg_return = returns.mean()
    std_dev = returns.std()
    worst_loss = returns.min()
    var_95 = returns.quantile(0.05)
    confidence_score, confidence_label = classify_confidence(sample_size, probability_positive, std_dev)
    return {
        "Sample Size": sample_size,
        "Probability Positive %": probability_positive,
        "Average Return %": avg_return,
        "Median Return %": returns.median(),
        "Best Historical Gain %": returns.max(),
        "Worst Historical Loss %": worst_loss,
        "Average Max Upside %": matches[upside_col].mean(),
        "Average Max Drawdown %": matches[drawdown_col].mean(),
        "VaR 95% %": var_95,
        "Return Std Dev %": std_dev,
        "Confidence Score": confidence_score,
        "Confidence Label": confidence_label,
        "Risk Label": classify_risk(worst_loss, var_95, std_dev),
        "Outlook": classify_outlook(probability_positive, avg_return),
    }


def build_probability_engine(observations: pd.DataFrame) -> pd.DataFrame:
    rows = []
    if observations.empty:
        return pd.DataFrame()

    for stock, group in observations.groupby("Stock"):
        frame = group.sort_values("Date").copy()
        current = frame.iloc[-1]
        history = frame.iloc[:-15].copy()
        if history.empty:
            continue

        strict_matches = history[
            history["Volume Ratio"].between(current["Volume Ratio"] - 0.30, current["Volume Ratio"] + 0.30)
            & history["Daily Return %"].between(current["Daily Return %"] - 1.50, current["Daily Return %"] + 1.50)
            & history["RSI"].between(current["RSI"] - 8, current["RSI"] + 8)
            & history["Trend Regime"].eq(current["Trend Regime"])
        ]
        broad_matches = history[
            history["Volume Regime"].eq(current["Volume Regime"])
            & history["Trend Regime"].eq(current["Trend Regime"])
            & history["RSI"].between(current["RSI"] - 12, current["RSI"] + 12)
        ]
        matches = strict_matches if len(strict_matches) >= 20 else broad_matches
        match_type = "Strict Similarity" if len(strict_matches) >= 20 else "Broad Regime Similarity"

        summaries = {horizon: summarize_matches(matches, horizon) for horizon in [5, 10, 15]}
        primary = summaries[10]
        rows.append(
            {
                "Stock": stock,
                "Latest Date": current["Date"],
                "Current Close": current["Close"],
                "Current Daily Return %": current["Daily Return %"],
                "Current 20D Return %": current["20D Return %"],
                "Current Volume Ratio": current["Volume Ratio"],
                "Current Volume Regime": current["Volume Regime"],
                "Current RSI": current["RSI"],
                "Current Trend Regime": current["Trend Regime"],
                "Current Relative Strength %": current["Relative Strength vs Nifty %"],
                "Price-Volume Behaviour": current["Price-Volume Behaviour"],
                "Match Type": match_type,
                "Matched Historical Cases": primary["Sample Size"],
                "5D Positive Probability %": summaries[5]["Probability Positive %"],
                "5D Average Return %": summaries[5]["Average Return %"],
                "5D Expected Range": f"{summaries[5]['Worst Historical Loss %']:.2f}% to {summaries[5]['Best Historical Gain %']:.2f}%",
                "10D Positive Probability %": summaries[10]["Probability Positive %"],
                "10D Average Return %": summaries[10]["Average Return %"],
                "10D Median Return %": summaries[10]["Median Return %"],
                "10D Best Historical Gain %": summaries[10]["Best Historical Gain %"],
                "10D Worst Historical Loss %": summaries[10]["Worst Historical Loss %"],
                "10D Average Max Drawdown %": summaries[10]["Average Max Drawdown %"],
                "10D VaR 95% %": summaries[10]["VaR 95% %"],
                "10D Confidence Score": summaries[10]["Confidence Score"],
                "10D Confidence Label": summaries[10]["Confidence Label"],
                "10D Risk Label": summaries[10]["Risk Label"],
                "15D Positive Probability %": summaries[15]["Probability Positive %"],
                "15D Average Return %": summaries[15]["Average Return %"],
                "Historical Outlook": summaries[10]["Outlook"],
                "Evidence Summary": (
                    f"Over the available history, {int(primary['Sample Size'])} similar {match_type.lower()} cases "
                    f"were found for {stock}. The 10D positive probability was "
                    f"{primary['Probability Positive %']:.1f}% with an average return of "
                    f"{primary['Average Return %']:.2f}% and average max drawdown of "
                    f"{primary['Average Max Drawdown %']:.2f}%."
                )
                if primary["Sample Size"]
                else "Insufficient historical evidence for this setup.",
            }
        )

    probability = pd.DataFrame(rows)
    if probability.empty:
        return probability
    return probability.sort_values(
        ["10D Confidence Score", "10D Positive Probability %", "10D Average Return %"],
        ascending=[False, False, False],
    )


def build_price_volume_matrix(observations: pd.DataFrame) -> pd.DataFrame:
    if observations.empty:
        return pd.DataFrame()
    matrix = (
        observations.groupby("Price-Volume Behaviour", as_index=False)
        .agg(
            Occurrences=("Stock", "count"),
            **{
                "5D Positive Probability %": ("Future 5D Return %", lambda x: x.gt(0).mean() * 100),
                "5D Average Return %": ("Future 5D Return %", "mean"),
                "10D Positive Probability %": ("Future 10D Return %", lambda x: x.gt(0).mean() * 100),
                "10D Average Return %": ("Future 10D Return %", "mean"),
                "15D Positive Probability %": ("Future 15D Return %", lambda x: x.gt(0).mean() * 100),
                "15D Average Return %": ("Future 15D Return %", "mean"),
            },
        )
        .sort_values("10D Positive Probability %", ascending=False)
    )
    return matrix


def build_validation_report(probability_outlook: pd.DataFrame) -> pd.DataFrame:
    if probability_outlook.empty:
        return pd.DataFrame()
    sample = probability_outlook["Matched Historical Cases"]
    confidence = probability_outlook["10D Confidence Score"]
    consistency = (probability_outlook["10D Positive Probability %"] - 50).abs()
    return pd.DataFrame(
        [
            {
                "Validation Check": "Historical sample support",
                "Result": f"{sample.median():.0f} median matched cases per stock",
                "Status": "Pass" if sample.median() >= 30 else "Weak",
            },
            {
                "Validation Check": "Confidence stability",
                "Result": f"{confidence.mean():.1f}/100 average confidence score",
                "Status": "Pass" if confidence.mean() >= 50 else "Weak",
            },
            {
                "Validation Check": "Outcome consistency",
                "Result": f"{consistency.mean():.1f} average distance from random 50% baseline",
                "Status": "Pass" if consistency.mean() >= 5 else "Weak",
            },
            {
                "Validation Check": "Cross-stock coverage",
                "Result": f"{probability_outlook['Stock'].nunique()} stocks evaluated",
                "Status": "Pass" if probability_outlook["Stock"].nunique() >= 50 else "Weak",
            },
            {
                "Validation Check": "Out-of-sample warning",
                "Result": "Current version reports historical evidence; formal train/test validation is a future extension.",
                "Status": "Research Note",
            },
        ]
    )


def dataframe_to_excel(
    price_data: pd.DataFrame,
    bullish: pd.DataFrame,
    bearish: pd.DataFrame,
    summary: pd.DataFrame,
    trend_analysis: pd.DataFrame,
    volume_intelligence: pd.DataFrame,
    volume_spike_reactions: pd.DataFrame,
    volume_profile: pd.DataFrame,
    probability_outlook: pd.DataFrame,
    price_volume_matrix: pd.DataFrame,
    validation_report: pd.DataFrame,
    rankings: dict[str, pd.DataFrame],
    threshold_tests: pd.DataFrame,
    backtest_trades: pd.DataFrame,
    backtest_metrics: pd.DataFrame,
    conclusion_points: list[str],
    threshold_pct: float,
    universe: str,
    analysis_years: int,
    lookback_days: int,
    require_same_direction: bool,
) -> bytes:
    output = BytesIO()
    dashboard = pd.DataFrame(
        {
            "Conclusion": conclusion_points,
        }
    )
    methodology = pd.DataFrame(
        {
            "Item": [
                "Universe",
                "Period",
                "Trend lookback",
                "Volume intelligence",
                "Historical probability engine",
                "Bullish signal",
                "Bearish signal",
                "Forward returns",
                "Direction filter",
                "Benchmark",
                "Note",
            ],
            "Value": [
                f"{universe} stocks",
                f"{analysis_years} year daily OHLCV data",
                f"{lookback_days} trading days",
                "Price-volume behaviour, volume spikes, pressure indices, accumulation/distribution, liquidity, and conviction scoring",
                "Current setup is matched against similar historical price-volume conditions to estimate 5D, 10D, and 15D outcome probabilities",
                f"3-day return >= +{threshold_pct:.2f}%",
                f"3-day return <= -{threshold_pct:.2f}%",
                "Next 3 trading days and next 5 trading days",
                "Required" if require_same_direction else "Not required",
                "Nifty 50 index (^NSEI), when available",
                "Exploratory analysis, not investment advice.",
            ],
        }
    )
    limitations = pd.DataFrame({"Limitations": LIMITATIONS})

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        dashboard.to_excel(writer, sheet_name="Dashboard", index=False)
        methodology.to_excel(writer, sheet_name="Methodology", index=False)
        summary.to_excel(writer, sheet_name="Stock Summary", index=False)
        trend_analysis.to_excel(writer, sheet_name="Trend Analysis", index=False)
        volume_intelligence.to_excel(writer, sheet_name="Volume Intelligence", index=False)
        volume_spike_reactions.to_excel(writer, sheet_name="Volume Spike Reactions", index=False)
        volume_profile.to_excel(writer, sheet_name="Volume Profile", index=False)
        probability_outlook.to_excel(writer, sheet_name="Probability Outlook", index=False)
        price_volume_matrix.to_excel(writer, sheet_name="PV Behaviour Matrix", index=False)
        validation_report.to_excel(writer, sheet_name="Validation Report", index=False)
        for name, frame in rankings.items():
            frame.to_excel(writer, sheet_name=name[:31], index=False)
        threshold_tests.to_excel(writer, sheet_name="Threshold Tests", index=False)
        backtest_metrics.to_excel(writer, sheet_name="Backtest Metrics", index=False)
        backtest_trades.to_excel(writer, sheet_name="Backtest Trades", index=False)
        bullish.to_excel(writer, sheet_name="Bullish Signals", index=False)
        bearish.to_excel(writer, sheet_name="Bearish Signals", index=False)
        limitations.to_excel(writer, sheet_name="Limitations", index=False)
        price_data.to_excel(writer, sheet_name="Price Data", index=False)

        header_fill = PatternFill("solid", fgColor="1F4E78")
        header_font = Font(color="FFFFFF", bold=True)
        positive_fill = PatternFill("solid", fgColor="D9EAD3")
        negative_fill = PatternFill("solid", fgColor="F4CCCC")
        for sheet in writer.book.worksheets:
            sheet.freeze_panes = "A2"
            sheet.auto_filter.ref = sheet.dimensions
            for cell in sheet[1]:
                cell.fill = header_fill
                cell.font = header_font
            for column_cells in sheet.columns:
                max_length = max(len(str(cell.value or "")) for cell in column_cells)
                sheet.column_dimensions[column_cells[0].column_letter].width = min(
                    max(max_length + 2, 12), 36
                )
                header = str(column_cells[0].value or "")
                if "%" in header and sheet.max_row > 1:
                    letter = column_cells[0].column_letter
                    range_ref = f"{letter}2:{letter}{sheet.max_row}"
                    sheet.conditional_formatting.add(
                        range_ref,
                        CellIsRule(
                            operator="greaterThan",
                            formula=["0"],
                            fill=positive_fill,
                        ),
                    )
                    sheet.conditional_formatting.add(
                        range_ref,
                        CellIsRule(
                            operator="lessThan",
                            formula=["0"],
                            fill=negative_fill,
                        ),
                    )

    return output.getvalue()


def build_rankings(summary: pd.DataFrame, minimum_signals: int) -> dict[str, pd.DataFrame]:
    if summary.empty:
        return {}

    bull_ready = summary[summary["Bullish Signal Count"].ge(minimum_signals)].copy()
    bear_ready = summary[summary["Bearish Signal Count"].ge(minimum_signals)].copy()
    rankings = {
        "Top Bullish Continuation": bull_ready.sort_values(
            ["Avg Next 5D After Bullish %", "Bullish Continuation Rate %"],
            ascending=[False, False],
        ).head(10),
        "Top Bullish Reversal": bull_ready.sort_values(
            ["Avg Next 5D After Bullish %", "Bullish Continuation Rate %"],
            ascending=[True, True],
        ).head(10),
        "Top Bearish Bounce": bear_ready.sort_values(
            ["Avg Next 5D After Bearish %", "Bearish Bounce Rate %"],
            ascending=[False, False],
        ).head(10),
        "Top Bearish Continuation": bear_ready.sort_values(
            ["Avg Next 5D After Bearish %", "Bearish Bounce Rate %"],
            ascending=[True, True],
        ).head(10),
    }
    return {name: frame for name, frame in rankings.items() if not frame.empty}


def run_threshold_tests(
    price_data: pd.DataFrame,
    benchmark_data: pd.DataFrame,
    require_same_direction: bool,
) -> pd.DataFrame:
    rows = []
    for threshold in THRESHOLD_TESTS:
        bullish, bearish, _summary = analyze_streaks(
            price_data,
            threshold_pct=threshold,
            require_same_direction=require_same_direction,
            benchmark_data=benchmark_data,
        )
        rows.append(
            {
                "Threshold %": threshold,
                "Bullish Signals": len(bullish),
                "Avg Bullish Next 5D %": bullish["Next 5D Return %"].mean()
                if not bullish.empty
                else np.nan,
                "Bullish Continuation Rate %": bullish["Next 5D Return %"].gt(0).mean() * 100
                if not bullish.empty
                else np.nan,
                "Bearish Signals": len(bearish),
                "Avg Bearish Next 5D %": bearish["Next 5D Return %"].mean()
                if not bearish.empty
                else np.nan,
                "Bearish Bounce Rate %": bearish["Next 5D Return %"].gt(0).mean() * 100
                if not bearish.empty
                else np.nan,
            }
        )
    return pd.DataFrame(rows)


def build_conclusion(
    bullish: pd.DataFrame,
    bearish: pd.DataFrame,
    summary: pd.DataFrame,
    minimum_signals: int,
) -> list[str]:
    points = []

    if not bullish.empty:
        bull_avg = bullish["Next 5D Return %"].mean()
        bull_rate = bullish["Next 5D Return %"].gt(0).mean() * 100
        points.append(
            f"After bullish 3-day streaks, the average next 5-day return was {bull_avg:.2f}% with a {bull_rate:.1f}% continuation rate."
        )

    if not bearish.empty:
        bear_avg = bearish["Next 5D Return %"].mean()
        bear_rate = bearish["Next 5D Return %"].gt(0).mean() * 100
        points.append(
            f"After bearish 3-day streaks, the average next 5-day return was {bear_avg:.2f}% with a {bear_rate:.1f}% bounce rate."
        )

    rankings = build_rankings(summary, minimum_signals)
    if "Top Bullish Continuation" in rankings:
        row = rankings["Top Bullish Continuation"].iloc[0]
        points.append(
            f"{row['Stock']} ranked strongest after bullish streaks among stocks with at least {minimum_signals} signals."
        )
    if "Top Bearish Bounce" in rankings:
        row = rankings["Top Bearish Bounce"].iloc[0]
        points.append(
            f"{row['Stock']} showed the strongest average bounce after bearish streaks among stocks with at least {minimum_signals} signals."
        )

    if "5D Alpha vs Nifty %" in bullish.columns or "5D Alpha vs Nifty %" in bearish.columns:
        points.append(
            "Benchmark comparison is included through 5-day alpha versus the Nifty 50 index."
        )

    points.append(
        "The result should be treated as exploratory statistical evidence, not as a trading recommendation."
    )
    return points


def filter_outputs(
    summary: pd.DataFrame,
    bullish: pd.DataFrame,
    bearish: pd.DataFrame,
    minimum_signals: int,
    result_filter: str,
) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    if summary.empty:
        return summary, bullish.copy(), bearish.copy()

    filtered_summary = summary[
        summary["Bullish Signal Count"].add(summary["Bearish Signal Count"]).ge(minimum_signals)
    ].copy()

    filtered_bullish = bullish.copy()
    filtered_bearish = bearish.copy()
    if result_filter != "All":
        filtered_bullish = filtered_bullish[filtered_bullish["Result"].eq(result_filter)]
        filtered_bearish = filtered_bearish[filtered_bearish["Result"].eq(result_filter)]

    return filtered_summary, filtered_bullish, filtered_bearish


def build_backtest_trades(
    bullish: pd.DataFrame,
    bearish: pd.DataFrame,
    strategy: str,
    transaction_cost_bps: float,
    min_alpha_pct: float,
) -> pd.DataFrame:
    trade_frames = []
    cost_pct = transaction_cost_bps / 100

    if strategy in {"Bullish continuation long", "Combined long signals"}:
        trades = bullish.copy()
        if not trades.empty:
            trades["Strategy Leg"] = "Bullish Long"
            trades["Direction"] = "Long"
            trades["Gross Trade Return %"] = trades["Next 5D Return %"]
            trade_frames.append(trades)

    if strategy in {"Bearish bounce long", "Combined long signals"}:
        trades = bearish.copy()
        if not trades.empty:
            trades["Strategy Leg"] = "Bearish Bounce Long"
            trades["Direction"] = "Long"
            trades["Gross Trade Return %"] = trades["Next 5D Return %"]
            trade_frames.append(trades)

    if strategy == "Bearish continuation short":
        trades = bearish.copy()
        if not trades.empty:
            trades["Strategy Leg"] = "Bearish Short"
            trades["Direction"] = "Short"
            trades["Gross Trade Return %"] = -trades["Next 5D Return %"]
            trade_frames.append(trades)

    if not trade_frames:
        return pd.DataFrame()

    all_trades = pd.concat(trade_frames, ignore_index=True)
    if min_alpha_pct and "5D Alpha vs Nifty %" in all_trades.columns:
        all_trades = all_trades[all_trades["5D Alpha vs Nifty %"].ge(min_alpha_pct)].copy()

    if all_trades.empty:
        return pd.DataFrame()

    all_trades["Net Trade Return %"] = all_trades["Gross Trade Return %"] - cost_pct
    all_trades["Win"] = all_trades["Net Trade Return %"].gt(0)
    all_trades = all_trades.sort_values(["Date", "Stock", "Strategy Leg"]).reset_index(drop=True)
    all_trades["Equity Curve"] = (1 + all_trades["Net Trade Return %"] / 100).cumprod()
    all_trades["Cumulative Return %"] = (all_trades["Equity Curve"] - 1) * 100
    all_trades["Drawdown %"] = (
        all_trades["Equity Curve"] / all_trades["Equity Curve"].cummax() - 1
    ) * 100

    output_columns = [
        "Date",
        "Stock",
        "Strategy Leg",
        "Direction",
        "Signal Type",
        "3-Day Return %",
        "Next 5D Return %",
        "Gross Trade Return %",
        "Net Trade Return %",
        "Win",
        "5D Alpha vs Nifty %",
        "Cumulative Return %",
        "Drawdown %",
    ]
    existing_columns = [column for column in output_columns if column in all_trades.columns]
    return all_trades[existing_columns]


def build_backtest_metrics(trades: pd.DataFrame, strategy: str) -> pd.DataFrame:
    if trades.empty:
        return pd.DataFrame(
            [
                {
                    "Metric": "Strategy",
                    "Value": strategy,
                },
                {
                    "Metric": "Trades",
                    "Value": 0,
                },
            ]
        )

    returns = trades["Net Trade Return %"]
    wins = returns[returns.gt(0)]
    losses = returns[returns.lt(0)]
    std = returns.std()
    sharpe_like = (returns.mean() / std) * np.sqrt(252 / 5) if std and not np.isnan(std) else np.nan
    profit_factor = wins.sum() / abs(losses.sum()) if abs(losses.sum()) > 0 else np.nan
    avg_alpha = (
        trades["5D Alpha vs Nifty %"].mean()
        if "5D Alpha vs Nifty %" in trades.columns
        else np.nan
    )

    metrics = [
        ("Strategy", strategy),
        ("Trades", len(trades)),
        ("Average Net Trade Return %", returns.mean()),
        ("Median Net Trade Return %", returns.median()),
        ("Win Rate %", trades["Win"].mean() * 100),
        ("Best Trade %", returns.max()),
        ("Worst Trade %", returns.min()),
        ("Return Std Dev %", std),
        ("Cumulative Return %", trades["Cumulative Return %"].iloc[-1]),
        ("Maximum Drawdown %", trades["Drawdown %"].min()),
        ("Sharpe-like Score", sharpe_like),
        ("Profit Factor", profit_factor),
        ("Average 5D Alpha vs Nifty %", avg_alpha),
    ]
    return pd.DataFrame({"Metric": [item[0] for item in metrics], "Value": [item[1] for item in metrics]})


def format_percent_columns(frame: pd.DataFrame) -> pd.DataFrame:
    formatted = frame.copy()
    percent_columns = [column for column in formatted.columns if "%" in column]
    for column in percent_columns:
        formatted[column] = formatted[column].round(2)
    return formatted


st.title("Quantitative Market Analysis Platform")
st.caption(
    "Beyond Prediction, A Framework for Systematizing Investment Research"
)

st.markdown(
    """
    <style>
    div[data-testid="stMetric"] {
        border: 1px solid #30363d;
        border-radius: 8px;
        padding: 12px;
        background: #161b22;
        color: #f8fafc;
    }
    div[data-testid="stMetric"] label,
    div[data-testid="stMetric"] [data-testid="stMetricLabel"] {
        color: #cbd5e1;
    }
    div[data-testid="stMetric"] [data-testid="stMetricValue"] {
        color: #ffffff;
        font-weight: 700;
    }
    .stTabs [data-baseweb="tab-list"] {
        gap: 8px;
    }
    </style>
    """,
    unsafe_allow_html=True,
)

selected_symbols: list[str] = []

with st.sidebar:
    st.header("Inputs")
    universe = st.selectbox(
        "Universe",
        list(INDEX_CONFIGS),
        index=1,
        help="Nifty 500 is the upgraded research universe; Nifty 50 remains available for faster testing.",
    )
    data_source = st.radio(
        "Data source",
        ["Fetch latest data", "Use cached data", "Upload my own data"],
        index=0,
    )
    analysis_years = st.slider(
        "Historical period (years)",
        min_value=1,
        max_value=5,
        value=5,
        step=1,
    )
    lookback_days = st.slider(
        "Trend lookback window",
        min_value=5,
        max_value=90,
        value=15,
        step=5,
        help="Used for trend strength, volatility, volume growth, RSI, and relative strength.",
    )
    volume_spike_multiplier = st.slider(
        "Volume spike threshold",
        min_value=1.2,
        max_value=3.0,
        value=1.5,
        step=0.1,
        help="Flags days where volume is this multiple of the rolling average volume.",
    )
    threshold_pct = st.slider(
        "3-day move threshold",
        min_value=1.0,
        max_value=10.0,
        value=5.0,
        step=0.5,
        help="Bullish signals are above this value; bearish signals are below the negative value.",
    )
    require_same_direction = st.checkbox(
        "Require all 3 days to move in the same direction",
        value=True,
    )
    minimum_signals = st.number_input(
        "Minimum signals for rankings",
        min_value=1,
        max_value=20,
        value=2,
        step=1,
    )
    result_filter = st.selectbox(
        "Signal table filter",
        [
            "All",
            "Continued Up",
            "Reversed Down",
            "Reversed Up / Bounce",
            "Continued Down",
        ],
    )
    st.divider()
    st.subheader("Backtest")
    backtest_strategy = st.selectbox(
        "Strategy",
        list(BACKTEST_STRATEGIES),
        help="Simple event-level simulation using the next 5-day return after each signal.",
    )
    transaction_cost_bps = st.number_input(
        "Transaction cost per trade (bps)",
        min_value=0.0,
        max_value=100.0,
        value=10.0,
        step=1.0,
    )
    min_alpha_pct = st.number_input(
        "Minimum 5D alpha filter %",
        min_value=-20.0,
        max_value=20.0,
        value=-20.0,
        step=0.5,
        help="Raise this to keep only trades that outperformed Nifty by at least this much.",
    )
    st.divider()
    end_date = st.date_input("End date", value=date.today())
    start_date = st.date_input("Start date", value=end_date - timedelta(days=365 * analysis_years))

    st.divider()
    st.caption("For uploaded files, use columns: Date, Stock, Open, High, Low, Close, Volume.")
    uploaded_file = st.file_uploader("Upload CSV or Excel", type=["csv", "xlsx", "xls"])

    if data_source == "Fetch latest data":
        constituents = load_index_symbols(universe)
        default_symbols = constituents["Yahoo Symbol"].dropna().astype(str).tolist()
        selected_symbols = st.multiselect(
            "Stocks",
            options=default_symbols,
            default=default_symbols,
            help="Uses Yahoo Finance tickers with .NS suffix.",
        )
    elif data_source == "Use cached data":
        st.caption(f"Cache file: {get_cache_path(universe)}")

    run_analysis = st.button("Run analysis", type="primary", use_container_width=True)


if start_date >= end_date:
    st.error("Start date must be earlier than end date.")
    st.stop()

if not run_analysis:
    if "analysis_payload" not in st.session_state:
        st.info(
            "Set the threshold and date range in the sidebar, then click Run analysis to build the tables and Excel report."
        )
        st.stop()
        raise SystemExit
    st.caption("Showing the last completed analysis. Click Run analysis to refresh.")

if run_analysis:
    cache_path = get_cache_path(universe)
    if data_source == "Fetch latest data":
        if not selected_symbols:
            st.warning("Select at least one stock.")
            st.stop()

        with st.spinner(f"Fetching {universe} price data. Large universes can take a few minutes..."):
            download_symbols = tuple(dict.fromkeys([*selected_symbols, BENCHMARK_SYMBOL]))
            downloaded_data = download_yfinance_data(download_symbols, start_date, end_date)
            price_data, benchmark_data = split_benchmark_data(downloaded_data)
            if not price_data.empty:
                save_cached_price_data(price_data, benchmark_data, cache_path)
    elif data_source == "Use cached data":
        cached_data = load_cached_price_data(cache_path)
        if cached_data.empty:
            st.warning("No cached data found yet. Fetch latest data once to create a cache.")
            st.stop()
        price_data, benchmark_data = split_benchmark_data(cached_data)
    else:
        if uploaded_file is None:
            st.info("Upload a CSV or Excel file to begin.")
            st.stop()
        try:
            uploaded_data = read_uploaded_data(uploaded_file)
            price_data, benchmark_data = split_benchmark_data(uploaded_data)
        except Exception as exc:
            st.error(str(exc))
            st.stop()

    if price_data.empty:
        st.error("No price data was available. Try a different date range or upload a file.")
        st.stop()

    with st.spinner("Calculating signals and forward returns..."):
        bullish_signals, bearish_signals, stock_summary = analyze_streaks(
            price_data,
            threshold_pct=threshold_pct,
            require_same_direction=require_same_direction,
            benchmark_data=benchmark_data,
        )
        threshold_tests = run_threshold_tests(
            price_data,
            benchmark_data,
            require_same_direction=require_same_direction,
        )
        trend_analysis = build_trend_analysis(
            price_data,
            benchmark_data,
            stock_summary,
            lookback_days=lookback_days,
        )
        volume_intelligence = build_volume_intelligence(
            price_data,
            trend_analysis,
            lookback_days=lookback_days,
        )
        volume_spike_reactions = build_volume_spike_reactions(
            price_data,
            lookback_days=lookback_days,
            spike_multiplier=volume_spike_multiplier,
        )
        volume_profile = build_long_term_volume_profile(price_data)
        historical_observations = build_historical_behavior_database(price_data, benchmark_data)
        probability_outlook = build_probability_engine(historical_observations)
        price_volume_matrix = build_price_volume_matrix(historical_observations)
        validation_report = build_validation_report(probability_outlook)
        rankings = build_rankings(stock_summary, minimum_signals)
        conclusion_points = build_conclusion(
            bullish_signals,
            bearish_signals,
            stock_summary,
            minimum_signals,
        )

    st.session_state["analysis_payload"] = {
        "price_data": price_data,
        "benchmark_data": benchmark_data,
        "bullish_signals": bullish_signals,
        "bearish_signals": bearish_signals,
        "stock_summary": stock_summary,
        "trend_analysis": trend_analysis,
        "volume_intelligence": volume_intelligence,
        "volume_spike_reactions": volume_spike_reactions,
        "volume_profile": volume_profile,
        "historical_observations": historical_observations,
        "probability_outlook": probability_outlook,
        "price_volume_matrix": price_volume_matrix,
        "validation_report": validation_report,
        "threshold_tests": threshold_tests,
        "rankings": rankings,
        "conclusion_points": conclusion_points,
        "threshold_pct": threshold_pct,
        "universe": universe,
        "analysis_years": analysis_years,
        "lookback_days": lookback_days,
        "volume_spike_multiplier": volume_spike_multiplier,
        "require_same_direction": require_same_direction,
        "minimum_signals": minimum_signals,
    }
else:
    payload = st.session_state["analysis_payload"]
    price_data = payload["price_data"]
    benchmark_data = payload.get("benchmark_data", pd.DataFrame())
    bullish_signals = payload["bullish_signals"]
    bearish_signals = payload["bearish_signals"]
    stock_summary = payload["stock_summary"]
    trend_analysis = payload.get("trend_analysis", pd.DataFrame())
    volume_intelligence = payload.get("volume_intelligence", pd.DataFrame())
    volume_spike_reactions = payload.get("volume_spike_reactions", pd.DataFrame())
    volume_profile = payload.get("volume_profile", pd.DataFrame())
    historical_observations = payload.get("historical_observations", pd.DataFrame())
    probability_outlook = payload.get("probability_outlook", pd.DataFrame())
    price_volume_matrix = payload.get("price_volume_matrix", pd.DataFrame())
    validation_report = payload.get("validation_report", pd.DataFrame())
    threshold_tests = payload.get("threshold_tests", pd.DataFrame())
    rankings = payload.get("rankings", build_rankings(stock_summary, minimum_signals))
    conclusion_points = payload.get(
        "conclusion_points",
        build_conclusion(bullish_signals, bearish_signals, stock_summary, minimum_signals),
    )
    threshold_pct = payload["threshold_pct"]
    universe = payload.get("universe", universe)
    analysis_years = payload.get("analysis_years", analysis_years)
    lookback_days = payload.get("lookback_days", lookback_days)
    volume_spike_multiplier = payload.get("volume_spike_multiplier", volume_spike_multiplier)
    require_same_direction = payload["require_same_direction"]
    minimum_signals = payload.get("minimum_signals", minimum_signals)

total_bullish = len(bullish_signals)
total_bearish = len(bearish_signals)
total_stocks = price_data["Stock"].nunique()
bullish_avg_5d = bullish_signals["Next 5D Return %"].mean() if total_bullish else np.nan
bearish_avg_5d = bearish_signals["Next 5D Return %"].mean() if total_bearish else np.nan
filtered_summary, filtered_bullish, filtered_bearish = filter_outputs(
    stock_summary,
    bullish_signals,
    bearish_signals,
    minimum_signals=minimum_signals,
    result_filter=result_filter,
)
backtest_trades = build_backtest_trades(
    bullish_signals,
    bearish_signals,
    strategy=backtest_strategy,
    transaction_cost_bps=transaction_cost_bps,
    min_alpha_pct=min_alpha_pct,
)
backtest_metrics = build_backtest_metrics(backtest_trades, backtest_strategy)

metric_cols = st.columns(6)
metric_cols[0].metric("Stocks Analyzed", total_stocks)
metric_cols[1].metric("Bullish Signals", total_bullish)
metric_cols[2].metric("Bearish Signals", total_bearish)
metric_cols[3].metric("Avg Bullish Next 5D", f"{bullish_avg_5d:.2f}%" if total_bullish else "NA")
metric_cols[4].metric("Avg Bearish Next 5D", f"{bearish_avg_5d:.2f}%" if total_bearish else "NA")
metric_cols[5].metric("Rows of Price Data", f"{len(price_data):,}")

if stock_summary.empty:
    st.warning(
        "No signals found with the current threshold. Try lowering the threshold or disabling the same-direction filter."
    )
    st.stop()

excel_bytes = dataframe_to_excel(
    price_data=price_data,
    bullish=bullish_signals,
    bearish=bearish_signals,
    summary=stock_summary,
    trend_analysis=trend_analysis,
    volume_intelligence=volume_intelligence,
    volume_spike_reactions=volume_spike_reactions,
    volume_profile=volume_profile,
    probability_outlook=probability_outlook,
    price_volume_matrix=price_volume_matrix,
    validation_report=validation_report,
    rankings=rankings,
    threshold_tests=threshold_tests,
    backtest_trades=backtest_trades,
    backtest_metrics=backtest_metrics,
    conclusion_points=conclusion_points,
    threshold_pct=threshold_pct,
    universe=universe,
    analysis_years=analysis_years,
    lookback_days=lookback_days,
    require_same_direction=require_same_direction,
)

st.download_button(
    "Download Excel Report",
    data=excel_bytes,
    file_name=f"{universe.lower().replace(' ', '')}_momentum_trend_analysis.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
)

tabs = st.tabs(
    [
        "Dashboard",
        "Trend Score",
        "Volume Intelligence",
        "Probability Engine",
        "Historical Matrix",
        "Validation",
        "Rankings",
        "Stock Summary",
        "Bullish Signals",
        "Bearish Signals",
        "Threshold Tests",
        "Backtest & Risk",
        "Charts",
        "Methodology",
        "Data",
    ]
)

with tabs[0]:
    st.subheader("Automatic Conclusion")
    for point in conclusion_points:
        st.write(f"- {point}")

    st.subheader("Presentation Snapshot")
    snapshot_cols = st.columns(3)
    snapshot_cols[0].metric("Current Threshold", f"{threshold_pct:.1f}%")
    snapshot_cols[1].metric("Ranking Signal Floor", minimum_signals)
    snapshot_cols[2].metric(
        "Benchmark",
        "Nifty 50 available" if not benchmark_data.empty else "Not available",
    )
    snapshot_cols = st.columns(3)
    snapshot_cols[0].metric("Universe", universe)
    snapshot_cols[1].metric("History", f"{analysis_years}Y")
    snapshot_cols[2].metric("Trend Lookback", f"{lookback_days}D")

with tabs[1]:
    st.subheader(f"{lookback_days}-Day Trend Score")
    st.caption(
        "Score = 30% lookback return + 20% continuation rate + 20% alpha + 10% low volatility + 10% volume growth + 10% relative strength."
    )
    if trend_analysis.empty:
        st.info("Trend analysis is unavailable for the selected data.")
    else:
        st.dataframe(format_percent_columns(trend_analysis), use_container_width=True, hide_index=True)
        trend_col_1, trend_col_2 = st.columns(2)
        with trend_col_1:
            fig = px.bar(
                trend_analysis.head(20),
                x="Stock",
                y="Trend Score",
                color="Trend Label",
                title=f"Top 20 {lookback_days}D Trend Scores",
            )
            st.plotly_chart(fig, use_container_width=True)
        with trend_col_2:
            return_col = f"{lookback_days}D Return %"
            relative_col = "Relative Strength vs Nifty %"
            fig = px.scatter(
                trend_analysis,
                x=return_col,
                y=relative_col,
                size="Trend Score",
                color="Trend Label",
                hover_name="Stock",
                title="Trend Return vs Relative Strength",
            )
            st.plotly_chart(fig, use_container_width=True)

with tabs[2]:
    st.subheader("Volume Intelligence & Price-Volume Behaviour")
    st.caption(
        "This module estimates buying pressure, selling pressure, accumulation/distribution behaviour, volume spikes, liquidity, and conviction using daily OHLCV data."
    )
    if volume_intelligence.empty:
        st.info("Volume intelligence is unavailable for the selected data.")
    else:
        high_volume_count = volume_intelligence["Volume Spike Category"].isin(
            ["High", "Very High", "Exceptional"]
        ).sum()
        volume_cols = st.columns(4)
        volume_cols[0].metric("High Volume Stocks", int(high_volume_count))
        volume_cols[1].metric(
            "Avg Conviction",
            f"{volume_intelligence['Conviction Score'].mean():.1f}/100",
        )
        volume_cols[2].metric(
            "Avg Buying Pressure",
            f"{volume_intelligence['Buying Pressure Index'].mean():.1f}/100",
        )
        volume_cols[3].metric(
            "Avg Selling Pressure",
            f"{volume_intelligence['Selling Pressure Index'].mean():.1f}/100",
        )

        st.dataframe(format_percent_columns(volume_intelligence), use_container_width=True, hide_index=True)
        chart_col_1, chart_col_2 = st.columns(2)
        with chart_col_1:
            fig = px.bar(
                volume_intelligence.head(20),
                x="Stock",
                y="Conviction Score",
                color="Price-Volume Interpretation",
                title="Top 20 Volume + Trend Conviction Scores",
            )
            st.plotly_chart(fig, use_container_width=True)
        with chart_col_2:
            fig = px.scatter(
                volume_intelligence,
                x="Buying Pressure Index",
                y="Selling Pressure Index",
                size="Volume Spike Ratio",
                color="Volume Spike Category",
                hover_name="Stock",
                title="Buying Pressure vs Selling Pressure",
            )
            st.plotly_chart(fig, use_container_width=True)

    st.subheader("Volume Spike Reactions")
    st.caption(
        f"Spike days are days where volume is at least {volume_spike_multiplier:.1f}x the prior rolling average volume."
    )
    if volume_spike_reactions.empty:
        st.info("No historical volume spike events matched the selected threshold.")
    else:
        st.dataframe(format_percent_columns(volume_spike_reactions), use_container_width=True, hide_index=True)
        spike_summary = (
            volume_spike_reactions.groupby("Spike Category", as_index=False)
            .agg(
                Events=("Stock", "count"),
                **{
                    "Avg Next 1D Return %": ("Next 1D Return %", "mean"),
                    "Avg Next 3D Return %": ("Next 3D Return %", "mean"),
                    "Avg Next 5D Return %": ("Next 5D Return %", "mean"),
                },
            )
            .sort_values("Events", ascending=False)
        )
        st.markdown("**Average Return After Volume Spike**")
        st.dataframe(format_percent_columns(spike_summary), use_container_width=True, hide_index=True)

with tabs[3]:
    st.subheader("Historical Behaviour Probability Engine")
    st.caption(
        "This matches each stock's current price-volume setup with similar historical conditions, then reports what happened over the next 5, 10, and 15 trading days."
    )
    if probability_outlook.empty:
        st.info("Probability outlook is unavailable. Use a longer history, ideally 5 years, to build enough historical cases.")
    else:
        prob_cols = st.columns(5)
        prob_cols[0].metric("Stocks Evaluated", probability_outlook["Stock"].nunique())
        prob_cols[1].metric(
            "Median Cases",
            f"{probability_outlook['Matched Historical Cases'].median():.0f}",
        )
        prob_cols[2].metric(
            "Avg 10D Probability",
            f"{probability_outlook['10D Positive Probability %'].mean():.1f}%",
        )
        prob_cols[3].metric(
            "Avg 10D Return",
            f"{probability_outlook['10D Average Return %'].mean():.2f}%",
        )
        prob_cols[4].metric(
            "Avg Confidence",
            f"{probability_outlook['10D Confidence Score'].mean():.1f}/100",
        )
        st.dataframe(format_percent_columns(probability_outlook), use_container_width=True, hide_index=True)

        chart_col_1, chart_col_2 = st.columns(2)
        with chart_col_1:
            fig = px.scatter(
                probability_outlook,
                x="10D Positive Probability %",
                y="10D Average Return %",
                size="Matched Historical Cases",
                color="Historical Outlook",
                hover_name="Stock",
                title="10D Probability vs Expected Return",
            )
            st.plotly_chart(fig, use_container_width=True)
        with chart_col_2:
            fig = px.bar(
                probability_outlook.head(20),
                x="Stock",
                y="10D Confidence Score",
                color="10D Confidence Label",
                title="Top 20 Historical Evidence Confidence Scores",
            )
            st.plotly_chart(fig, use_container_width=True)

        st.subheader("Explainability")
        for explanation in probability_outlook.head(5)["Evidence Summary"]:
            st.write(f"- {explanation}")

with tabs[4]:
    st.subheader("Historical Price-Volume Behaviour Matrix")
    if price_volume_matrix.empty:
        st.info("Historical matrix is unavailable for the selected data.")
    else:
        st.dataframe(format_percent_columns(price_volume_matrix), use_container_width=True, hide_index=True)
        fig = px.bar(
            price_volume_matrix,
            x="Price-Volume Behaviour",
            y="10D Positive Probability %",
            color="10D Average Return %",
            title="Historical Outcomes by Price-Volume Behaviour",
            color_continuous_scale="RdYlGn",
        )
        st.plotly_chart(fig, use_container_width=True)

    st.subheader("Long-Term Volume Profile")
    if volume_profile.empty:
        st.info("Volume profile is unavailable for the selected data.")
    else:
        st.dataframe(format_percent_columns(volume_profile), use_container_width=True, hide_index=True)

with tabs[5]:
    st.subheader("Validation Report")
    st.caption(
        "Before trusting a rule, this checks sample size, confidence stability, outcome consistency, cross-stock coverage, and the need for future out-of-sample validation."
    )
    if validation_report.empty:
        st.info("Validation report is unavailable because no probability outlook was generated.")
    else:
        st.dataframe(validation_report, use_container_width=True, hide_index=True)

with tabs[6]:
    st.subheader("Top 10 Ranking Tables")
    if not rankings:
        st.info("No ranking tables available with the current minimum signal count.")
    for name, frame in rankings.items():
        st.markdown(f"**{name}**")
        st.dataframe(format_percent_columns(frame), use_container_width=True, hide_index=True)

with tabs[7]:
    st.subheader("Stock-wise Performance Summary")
    st.dataframe(format_percent_columns(filtered_summary), use_container_width=True, hide_index=True)

with tabs[8]:
    st.subheader("Stocks After Strong 3-Day Up Moves")
    st.dataframe(format_percent_columns(filtered_bullish), use_container_width=True, hide_index=True)

with tabs[9]:
    st.subheader("Stocks After Strong 3-Day Down Moves")
    st.dataframe(format_percent_columns(filtered_bearish), use_container_width=True, hide_index=True)

with tabs[10]:
    st.subheader("Parameter Testing")
    st.caption("This checks whether the conclusion changes when the 3-day move threshold changes.")
    st.dataframe(format_percent_columns(threshold_tests), use_container_width=True, hide_index=True)
    if not threshold_tests.empty:
        fig = px.line(
            threshold_tests,
            x="Threshold %",
            y=["Avg Bullish Next 5D %", "Avg Bearish Next 5D %"],
            markers=True,
            title="Average Next 5D Return Across Thresholds",
        )
        st.plotly_chart(fig, use_container_width=True)

with tabs[11]:
    st.subheader("Event-Level Backtest and Risk")
    st.caption(BACKTEST_STRATEGIES[backtest_strategy])
    st.caption(
        "This is a simplified research backtest using signal-date forward returns. It does not model overlapping trades, margin, liquidity, stop-losses, taxes, or execution slippage beyond the cost input."
    )

    metric_lookup = dict(zip(backtest_metrics["Metric"], backtest_metrics["Value"]))
    risk_cols = st.columns(5)
    risk_cols[0].metric("Trades", int(metric_lookup.get("Trades", 0)))
    risk_cols[1].metric(
        "Avg Net Trade",
        f"{float(metric_lookup.get('Average Net Trade Return %', np.nan)):.2f}%"
        if "Average Net Trade Return %" in metric_lookup
        else "NA",
    )
    risk_cols[2].metric(
        "Win Rate",
        f"{float(metric_lookup.get('Win Rate %', np.nan)):.1f}%"
        if "Win Rate %" in metric_lookup
        else "NA",
    )
    risk_cols[3].metric(
        "Cumulative Return",
        f"{float(metric_lookup.get('Cumulative Return %', np.nan)):.2f}%"
        if "Cumulative Return %" in metric_lookup
        else "NA",
    )
    risk_cols[4].metric(
        "Max Drawdown",
        f"{float(metric_lookup.get('Maximum Drawdown %', np.nan)):.2f}%"
        if "Maximum Drawdown %" in metric_lookup
        else "NA",
    )

    st.dataframe(format_percent_columns(backtest_metrics), use_container_width=True, hide_index=True)
    if backtest_trades.empty:
        st.info("No trades matched the selected backtest settings.")
    else:
        chart_col_1, chart_col_2 = st.columns(2)
        with chart_col_1:
            fig = px.line(
                backtest_trades,
                x="Date",
                y="Cumulative Return %",
                title="Backtest Cumulative Return",
                markers=True,
            )
            st.plotly_chart(fig, use_container_width=True)
        with chart_col_2:
            fig = px.area(
                backtest_trades,
                x="Date",
                y="Drawdown %",
                title="Backtest Drawdown",
            )
            st.plotly_chart(fig, use_container_width=True)
        st.subheader("Backtest Trades")
        st.dataframe(format_percent_columns(backtest_trades), use_container_width=True, hide_index=True)

with tabs[12]:
    chart_col_1, chart_col_2 = st.columns(2)
    with chart_col_1:
        if not bullish_signals.empty:
            fig = px.histogram(
                bullish_signals,
                x="Next 5D Return %",
                nbins=30,
                title="Next 5D Returns After Bullish Signals",
                color_discrete_sequence=["#2f9e44"],
            )
            st.plotly_chart(fig, use_container_width=True)
    with chart_col_2:
        if not bearish_signals.empty:
            fig = px.histogram(
                bearish_signals,
                x="Next 5D Return %",
                nbins=30,
                title="Next 5D Returns After Bearish Signals",
                color_discrete_sequence=["#c92a2a"],
            )
            st.plotly_chart(fig, use_container_width=True)

    heatmap_source = stock_summary.melt(
        id_vars="Stock",
        value_vars=[
            "Avg Next 5D After Bullish %",
            "Avg Next 5D After Bearish %",
        ],
        var_name="Signal",
        value_name="Average Next 5D Return %",
    ).dropna()
    fig = px.density_heatmap(
        heatmap_source,
        x="Signal",
        y="Stock",
        z="Average Next 5D Return %",
        histfunc="avg",
        title="Average Next 5D Return by Stock and Signal Type",
        color_continuous_scale="RdYlGn",
    )
    st.plotly_chart(fig, use_container_width=True)

with tabs[13]:
    st.subheader("Methodology")
    st.write(
        f"Bullish signals are defined as 3-day returns greater than or equal to +{threshold_pct:.2f}%."
    )
    st.write(
        f"Bearish signals are defined as 3-day returns less than or equal to -{threshold_pct:.2f}%."
    )
    st.write("Forward performance is measured over the next 3 and 5 trading days.")
    st.write(
        "The backtest module converts signal-level forward returns into a simplified event-level trade simulation."
    )
    st.write(
        f"The trend score ranks stocks using a configurable {lookback_days}-trading-day lookback window."
    )
    st.write(
        "The volume intelligence module uses daily OHLCV data to infer pressure and participation. It does not claim exact institutional buying or selling volume."
    )
    st.write(
        "Price-volume matrix: price up with volume up is treated as accumulation, price up with volume down as weak rally, price down with volume up as distribution, and price down with volume down as weak selling."
    )
    st.write(
        "The conviction score combines trend score, volume intelligence score, relative strength, continuation rate, and benchmark alpha."
    )
    st.write(
        "The probability engine matches the current setup against historical observations with similar volume ratio, price movement, RSI, and trend regime, then summarizes 5D, 10D, and 15D outcomes."
    )
    st.write(
        "Confidence depends on matched sample size, consistency away from a random 50% baseline, and return stability."
    )
    st.write(
        "The same-direction filter is enabled."
        if require_same_direction
        else "The same-direction filter is disabled."
    )
    st.subheader("Limitations")
    for limitation in LIMITATIONS:
        st.write(f"- {limitation}")

with tabs[14]:
    st.subheader("Price Data Used")
    st.dataframe(price_data.sort_values(["Date", "Stock"]), use_container_width=True, hide_index=True)
    if not benchmark_data.empty:
        st.subheader("Benchmark Data Used")
        st.dataframe(
            benchmark_data.sort_values(["Date", "Stock"]),
            use_container_width=True,
            hide_index=True,
        )
